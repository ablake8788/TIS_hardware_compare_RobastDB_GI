##
# TIS_hardware_compare_orig.py


"""
Generate a Titanium vs competitor comparison Excel file using the OpenAI Responses API.

Setup:
  pip install openai openpyxl
  export OPENAI_API_KEY="your_api_key"

Usage:
  python generate_competitor_comparison.py \
      --competitor "PointCentral" \
      --url "https://www.pointcentral.com/" \
      --output "Titanium_vs_PointCentral_Comparison.xlsx"

Optional:
  python generate_competitor_comparison.py \
      --competitor "Kairos" \
      --url "https://www.kairoswater.io/" \
      --model "gpt-5.4" \
      --output "Titanium_vs_Kairos_Comparison.xlsx"
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from typing import List, Dict

from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


TITANIUM_HARDWARE = [
    {
        "hardware": "Smart Thermostat (LoRaWAN and EnOcean)",
        "description": "Smart thermostat for HVAC control using LoRaWAN and EnOcean communication options.",
    },
    {
        "hardware": "TRV (thermostatic radiator valve actuator)",
        "description": "Actuator for controlling radiator valves and room-level heating.",
    },
    {
        "hardware": "Pulse counter (gas, electric, and water)",
        "description": "Counts utility meter pulses for gas, electric, and water monitoring.",
    },
    {
        "hardware": "CT clamp",
        "description": "Current transformer clamp for measuring electrical current usage.",
    },
    {
        "hardware": "Relay (on/off)",
        "description": "Relay for basic remote on/off switching control.",
    },
    {
        "hardware": "Relay (on/off and dimming)",
        "description": "Relay for remote on/off control with dimming capability.",
    },
    {
        "hardware": "Dry contact relay",
        "description": "Relay with dry contact output for integration with external control circuits.",
    },
    {
        "hardware": "Rope (1-meter and 5-meter) water leak sensor",
        "description": "Cable-style leak detection sensor for identifying water presence over a length.",
    },
    {
        "hardware": "Probe (spot) water leak sensor",
        "description": "Spot leak sensor for localized water detection at a specific point.",
    },
    {
        "hardware": "Water valve control",
        "description": "Control device for remote opening and closing of water valves.",
    },
    {
        "hardware": "Water flow meter",
        "description": "Meter for monitoring water flow and usage.",
    },
    {
        "hardware": "Motion sensor",
        "description": "Sensor for detecting motion or occupancy in a space.",
    },
    {
        "hardware": "Door/window sensor",
        "description": "Sensor for detecting door or window open/close status.",
    },
    {
        "hardware": "Accurate people counter",
        "description": "Sensor for counting people entering or leaving an area.",
    },
    {
        "hardware": "Acceleration sensor",
        "description": "Sensor for measuring movement, tilt, or vibration.",
    },
    {
        "hardware": "Pipe temperature",
        "description": "Sensor for measuring pipe surface temperature.",
    },
    {
        "hardware": "Ambient temperature and humidity",
        "description": "Sensor for monitoring room temperature and humidity.",
    },
    {
        "hardware": "Temperature probe sensors (e.g., air ducts)",
        "description": "Probe sensors for measuring temperature in ducts and other targeted locations.",
    },
    {
        "hardware": "CO2 sensor",
        "description": "Sensor for measuring carbon dioxide levels.",
    },
    {
        "hardware": "VOC sensor",
        "description": "Sensor for measuring volatile organic compounds in the air.",
    },
    {
        "hardware": "Light (lux) sensor",
        "description": "Sensor for measuring ambient light levels.",
    },
    {
        "hardware": "Rocker switch (remote on/off and dimming)",
        "description": "Remote switch control for on/off and dimming functions.",
    },
    {
        "hardware": "Ultrasonic tank level sensor",
        "description": "Ultrasonic sensor for monitoring tank fill levels.",
    },
    {
        "hardware": "Gateway with cellular communication",
        "description": "Gateway device with cellular connectivity for network backhaul and device communication.",
    },
]

ALLOWED_STATUS = {"Yes", "Partial", "No"}


def build_prompt(competitor: str, url: str) -> str:
    hardware_lines = "\n".join(
        f'- {item["hardware"]} | {item["description"]}' for item in TITANIUM_HARDWARE
    )
    return f"""
You are creating a competitive hardware comparison for Titanium.

Competitor name: {competitor}
Competitor website: {url}

Titanium hardware list:
{hardware_lines}

Task:
Return ONLY valid JSON as an array of 24 objects.
Each object must have exactly these keys:
- "Titanium Hardware"
- "Description"
- "{competitor} Equivalent"
- "Comparable"

Rules:
- Use the Titanium hardware names exactly as provided.
- Keep Description concise and business-ready.
- For the competitor column, identify the closest real equivalent.
- If none exists, write exactly: "No direct {competitor} equivalent identified"
- "Comparable" must be one of: "Yes", "Partial", "No"
- Use "Yes" only for direct or strong equivalents.
- Use "Partial" for narrower, indirect, integration-based, or limited overlap.
- Use "No" for no meaningful equivalent.
- Be conservative and evidence-based.
- Emphasize Titanium as a broader multi-domain platform when the competitor is narrower.
- No markdown. No prose. No commentary. JSON only.
""".strip()


def extract_json(text: str) -> List[Dict[str, str]]:
    text = text.strip()

    # Try raw JSON first.
    try:
        data = json.loads(text)
        if isinstance(data, list):
            return data
    except json.JSONDecodeError:
        pass

    # Fallback: extract first JSON array from mixed text.
    match = re.search(r"(\[\s*\{.*\}\s*\])", text, re.DOTALL)
    if not match:
        raise ValueError("Could not find a JSON array in model output.")

    data = json.loads(match.group(1))
    if not isinstance(data, list):
        raise ValueError("Parsed JSON is not a list.")
    return data


def validate_rows(rows: List[Dict[str, str]], competitor: str) -> List[Dict[str, str]]:
    expected_keys = {
        "Titanium Hardware",
        "Description",
        f"{competitor} Equivalent",
        "Comparable",
    }

    by_name = {item["hardware"]: item["description"] for item in TITANIUM_HARDWARE}
    output: List[Dict[str, str]] = []

    for row in rows:
        if set(row.keys()) != expected_keys:
            raise ValueError(f"Unexpected keys in row: {row.keys()}")

        hardware = row["Titanium Hardware"]
        if hardware not in by_name:
            raise ValueError(f"Unknown Titanium hardware item: {hardware}")

        status = row["Comparable"]
        if status not in ALLOWED_STATUS:
            raise ValueError(f"Invalid Comparable status: {status}")

        # Normalize description to Titanium source description for consistency.
        output.append(
            {
                "Titanium Hardware": hardware,
                "Description": by_name[hardware],
                f"{competitor} Equivalent": row[f"{competitor} Equivalent"].strip(),
                "Comparable": status,
            }
        )

    # Ensure all 24 rows exist and appear in Titanium order.
    output_by_hw = {row["Titanium Hardware"]: row for row in output}
    missing = [item["hardware"] for item in TITANIUM_HARDWARE if item["hardware"] not in output_by_hw]
    if missing:
        raise ValueError(f"Missing hardware rows: {missing}")

    ordered = [output_by_hw[item["hardware"]] for item in TITANIUM_HARDWARE]
    return ordered


def generate_rows(client: OpenAI, competitor: str, url: str, model: str) -> List[Dict[str, str]]:
    prompt = build_prompt(competitor, url)

    response = client.responses.create(
        model=model,
        input=prompt,
    )

    rows = extract_json(response.output_text)
    return validate_rows(rows, competitor)


def save_excel(rows: List[Dict[str, str]], competitor: str, url: str, output_path: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Comparison"

    headers = ["Titanium Hardware", "Description", f"{competitor} Equivalent", "Comparable"]
    ws.append(headers)

    for row in rows:
        ws.append([
            row["Titanium Hardware"],
            row["Description"],
            row[f"{competitor} Equivalent"],
            row["Comparable"],
        ])

    header_fill = PatternFill(fill_type="solid", start_color="D9D9D9", end_color="D9D9D9")
    green_fill = PatternFill(fill_type="solid", start_color="C6E0B4", end_color="C6E0B4")
    yellow_fill = PatternFill(fill_type="solid", start_color="FFF2CC", end_color="FFF2CC")
    red_fill = PatternFill(fill_type="solid", start_color="F4CCCC", end_color="F4CCCC")
    thin_gray = Side(style="thin", color="BFBFBF")

    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=thin_gray)

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
        for idx, cell in enumerate(row, start=1):
            if idx != 4:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        status_cell = row[3]
        status_cell.font = Font(bold=True)
        status_cell.alignment = Alignment(horizontal="center", vertical="center")

        if status_cell.value == "Yes":
            status_cell.fill = green_fill
        elif status_cell.value == "Partial":
            status_cell.fill = yellow_fill
        else:
            status_cell.fill = red_fill

    ws.column_dimensions["A"].width = 42
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 56
    ws.column_dimensions["D"].width = 14
    ws.freeze_panes = "A2"
    ws.row_dimensions[1].height = 26

    # Add a basic sources sheet for traceability.
    src = wb.create_sheet("Sources")
    src.append(["Source", "URL"])
    src.append([f"{competitor} website", url])

    for cell in src[1]:
        cell.fill = header_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    src.column_dimensions["A"].width = 24
    src.column_dimensions["B"].width = 80

    wb.save(output_path)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Generate a Titanium competitor comparison Excel file.")
    parser.add_argument("--competitor", required=True, help="Competitor name, e.g. PointCentral")
    parser.add_argument("--url", required=True, help="Competitor website URL")
    parser.add_argument("--output", required=True, help="Output .xlsx path")
    parser.add_argument("--model", default="gpt-5.4", help="OpenAI model name")
    return parser.parse_args()


def main() -> int:
    args = parse_args()

    if not os.environ.get("OPENAI_API_KEY"):
        print("Error: OPENAI_API_KEY environment variable is not set.", file=sys.stderr)
        return 1

    client = OpenAI()

    try:
        rows = generate_rows(client, args.competitor, args.url, args.model)
        save_excel(rows, args.competitor, args.url, args.output)
    except Exception as exc:
        print(f"Failed: {exc}", file=sys.stderr)
        return 1

    print(f"Created: {args.output}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
